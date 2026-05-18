#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
respiria_cima_updater.py  —  v7.0 DEFINITIU
=============================================
Autora: Sílvia Álvarez Vega · ICS Atenció Primària Girona

MODES:
  --mode=detecta           Consulta CIMA, detecta CNs nous → afegeix a Excel
  --mode=comprova-pendents Compta "NOU — PENDENT" → pendents_count.txt
  --mode=comprova-publicar Compta novetats validades → publicar_count.txt
  --mode=regenera          Fusiona HTML actual + novetats validades → HTML nou
  --mode=tot               detecta + regenera

LÒGICA DETECTA:
  1. Consulta CIMA: vias=78 + atc=R03 + comerc=1
  2. Filtra per CN: si ja existeix al Excel → ignora
  3. Filtra formes no inhalatòries (comprimits, xarops, injectables, nebulitzadors)
  4. Infere tots els camps automàticament des de vtm.nombre i nom del dispositiu
  5. Afegeix a l'Excel com "NOU — PENDENT" per a revisió

LÒGICA REGENERA (FUSIÓ):
  - Conserva els fàrmacs ja publicats al HTML
  - Afegeix NOMÉS novetats amb col N buida + col M URL CIMA + col L data
"""

import argparse, json, os, re, sys, time
from datetime import date
import openpyxl, requests

# ── Configuració ──────────────────────────────────────────────────────────────
GITHUB_OWNER = "infermeriap-cpu"
GITHUB_REPO  = "respiria"
EXCEL_FNAME  = "inhaladors_MPOC_ASMA_Avanzado_FINAL.xlsx"
HTML_FNAME   = "RespirIA_v2.html"
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")

CIMA_URL    = "https://cima.aemps.es/cima/rest/medicamentos"
CIMA_PARAMS = {"vias": "78", "atc": "R03", "comerc": "1", "pagina": 1, "tamPagina": 100}

# ── Formes farmacèutiques exactes de la CIMA ─────────────────────────────────
# Formes que SÍ incorporem (noms exactes tal com retorna la CIMA)
FORMES_OK = {
    "SUSPENSIÓN PARA INHALACIÓN EN ENVASE A PRESIÓN",
    "SOLUCIÓN PARA INHALACIÓN EN ENVASE A PRESIÓN",
    "POLVO PARA INHALACIÓN",
    "POLVO PARA INHALACIÓN (UNIDOSIS)",
    "POLVO PARA INHALACIÓN (CÁPSULA DURA)",
    "POLVO PARA INHALACION",
    "POLVO PARA INHALACION (CAPSULA DURA)",
    "SOLUCIÓN PARA INHALACIÓN",
    "SOLUCIÓN PARA INHALACIÓN DEL VAPOR",
    "LÍQUIDO PARA INHALACIÓN DEL VAPOR",
    "SOLUCIÓN PARA INHALACION DEL VAPOR",
}

# Formes que NO incorporem (nebulitzadors)
FORMES_NO_INC = {
    "SOLUCIÓN PARA INHALACIÓN POR NEBULIZADOR",
    "SUSPENSIÓN PARA INHALACIÓN POR NEBULIZADOR",
    "POLVO PARA SOLUCIÓN PARA INHALACIÓN POR NEBULIZADOR",
    "POLVO Y DISOLVENTE PARA SOLUCIÓN PARA INHALACIÓN POR NEBULIZADOR",
    "SOLUCIÓN ORAL O CONCENTRADO PARA INHALACIÓN POR NEBULIZADOR",
}

def es_forma_inhalatoria(forma: str) -> bool:
    """Comprova si la forma farmacèutica és un inhalador vàlid (no nebulitzador)."""
    if not forma:
        return False
    forma_upper = forma.upper().strip()
    # Primer: si és nebulitzador → NO
    if forma_upper in FORMES_NO_INC:
        return False
    if "NEBULIZADOR" in forma_upper or "NEBULITZADOR" in forma_upper:
        return False
    # Si és una forma explícitament vàlida → SÍ
    if forma_upper in FORMES_OK:
        return True
    # Si conté "INHAL" i no és nebulitzador → SÍ (per cobrir variacions)
    if "INHAL" in forma_upper:
        return True
    # Resta → NO (comprimits, xarops, injectables, etc.)
    return False

# ── DISPOSITIU_MAP: clau → (tipus, co2, flux, maniobra, link_scientia) ─────────
# Ordre important: més específic primer
DISPOSITIU_MAP = [
    # IVS
    (["respimat"],                          "IVS",      "🟢", "20-30 l/m", "Lenta",   "https://scientiasalut.gencat.cat/handle/11351/11891"),
    # IPS uni
    (["breezhaler"],                        "IPS-uni",  "🟢", ">90 l/m",   "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11816"),
    (["aerolizer"],                         "IPS-uni",  "🟢", ">90 l/m",   "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11815"),
    (["handihaler"],                        "IPS-uni",  "🟢", "<50 l/m",   "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11879"),
    (["zonda"],                             "IPS-uni",  "🟢", "30-60 l/m", "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11895"),
    (["tavulus"],                           "IPS-uni",  "🟢", "<50 l/m",   "Ràpida",  ""),
    # IPS multi
    (["turbuhaler"],                        "IPS-multi","🟢", "50-60 l/m", "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11893"),
    (["accuhaler"],                         "IPS-multi","🟢", "60-90 l/m", "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11813"),
    (["genuair"],                           "IPS-multi","🟢", "60-90 l/m", "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11878"),
    (["nexthaler"],                         "IPS-multi","🟢", "60-90 l/m", "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11881"),
    (["spiromax"],                          "IPS-multi","🟢", "60-90 l/m", "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11892"),
    (["novolizer"],                         "IPS-multi","🟢", "60-90 l/m", "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11882"),
    (["ellipta"],                           "IPS-multi","🟢", "<50 l/m",   "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11818"),
    (["easyhaler"],                         "IPS-multi","🟢", "<50 l/m",   "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11817"),
    (["twisthaler"],                        "IPS-multi","🟢", "<50 l/m",   "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11894"),
    (["clickhaler"],                        "IPS-multi","🟢", "<50 l/m",   "Ràpida",  ""),
    (["forspiro"],                          "IPS-multi","🟢", ">90 l/m",   "Ràpida",  "https://scientiasalut.gencat.cat/handle/11351/11819"),
    # ICP — pressuritzats
    (["modulite","aerosphere","alvesco"],   "ICP",      "🔴", "20-30 l/m", "Lenta",   "https://scientiasalut.gencat.cat/handle/11351/11880"),
    (["suspension","suspensió","aerosol",
      "presurizado","pressuritzat","pmdi",
      "envase a presion","envàs a pressió"],"ICP",      "🔴", "20-30 l/m", "Lenta",   "https://scientiasalut.gencat.cat/handle/11351/11880"),
]

def inferir_dispositiu(nom_comercial: str, forma: str):
    """Retorna (tipus, co2, flux, maniobra, link) des del nom comercial i forma farmacèutica."""
    text = (nom_comercial + " " + forma).lower()
    for keywords, tipus, co2, flux, maniobra, link in DISPOSITIU_MAP:
        for kw in keywords:
            if kw in text:
                return tipus, co2, flux, maniobra, link
    # Per defecte: ICP
    return "ICP", "🔴", "20-30 l/m", "Lenta", "https://scientiasalut.gencat.cat/handle/11351/11880"

# ── VTM_MAP: vtm.nombre → (classe, dosi_mpoc, dosi_max, phf, matma) ──────────
VTM_MAP = {
    # SABA
    "salbutamol":                           ("SABA",         "100-200 mcg si cal",    "200 mcg c/6h",       True,  False),
    "terbutalina":                          ("SABA",         "500 mcg si cal",         "6.000 mcg/24h",     False, False),
    # SAMA
    "bromuro de ipratropio":                ("SAMA",         "40 mcg c/6-8h si cal",  "240 mcg/24h",        True,  False),
    "ipratropio":                           ("SAMA",         "40 mcg c/6-8h si cal",  "240 mcg/24h",        True,  False),
    # LABA
    "salmeterol":                           ("LABA",         "50 mcg c/12h",           "100 mcg c/12h",     True,  False),
    "formoterol":                           ("LABA",         "12 mcg c/12h",           "24 mcg c/12h",      True,  False),
    "formoterol fumarato":                  ("LABA",         "12 mcg c/12h",           "24 mcg c/12h",      True,  False),
    "indacaterol":                          ("LABA",         "150 mcg c/24h",          "300 mcg c/24h",     True,  False),
    "olodaterol":                           ("LABA",         "5 mcg c/24h",            "5 mcg c/24h",       False, False),
    # LAMA
    "tiotropio":                            ("LAMA",         "18 mcg c/24h",           "18 mcg c/24h",      True,  False),
    "bromuro de tiotropio":                 ("LAMA",         "18 mcg c/24h",           "18 mcg c/24h",      True,  False),
    "aclidinio":                            ("LAMA",         "322 mcg c/12h",          "322 mcg c/12h",     False, False),
    "bromuro de aclidinio":                 ("LAMA",         "322 mcg c/12h",          "322 mcg c/12h",     False, False),
    "glicopirronio":                        ("LAMA",         "44 mcg c/24h",           "44 mcg c/24h",      False, False),
    "bromuro de glicopirronio":             ("LAMA",         "44 mcg c/24h",           "44 mcg c/24h",      False, False),
    "umeclidinio":                          ("LAMA",         "55 mcg c/24h",           "55 mcg c/24h",      False, False),
    "bromuro de umeclidinio":               ("LAMA",         "55 mcg c/24h",           "55 mcg c/24h",      False, False),
    # GCI
    "budesonida":                           ("GCI",          "200-400 mcg c/12h",      "800 mcg c/12h",     False, False),
    "beclometasona":                        ("GCI",          "250-500 mcg c/12h",      "2.000 mcg/24h",     False, False),
    "dipropionato de beclometasona":        ("GCI",          "250-500 mcg c/12h",      "2.000 mcg/24h",     False, False),
    "fluticasona":                          ("GCI",          "250-500 mcg c/12h",      "500 mcg c/12h",     False, False),
    "propionato de fluticasona":            ("GCI",          "250-500 mcg c/12h",      "500 mcg c/12h",     False, False),
    "furoato de fluticasona":               ("GCI",          "92 mcg c/24h",           "184 mcg c/24h",     False, False),
    "ciclesonida":                          ("GCI",          "160 mcg c/24h",          "320 mcg c/12h",     False, False),
    "mometasona":                           ("GCI",          "200 mcg c/24h",          "400 mcg c/12h",     False, False),
    "furoato de mometasona":                ("GCI",          "200 mcg c/24h",          "400 mcg c/12h",     False, False),
    # LABA/LAMA
    "indacaterol + glicopirronio":          ("LABA/LAMA",    "85/43 mcg c/24h",        "85/43 mcg c/24h",   False, False),
    "aclidinio + formoterol":               ("LABA/LAMA",    "340/12 mcg c/12h",       "340/12 mcg c/12h",  False, False),
    "umeclidinio + vilanterol":             ("LABA/LAMA",    "55/22 mcg c/24h",        "55/22 mcg c/24h",   False, False),
    "tiotropio + olodaterol":               ("LABA/LAMA",    "5/5 mcg c/24h",          "5/5 mcg c/24h",     False, False),
    "glicopirronio + formoterol":           ("LABA/LAMA",    "14.4/10 mcg c/12h",      "14.4/10 mcg c/12h", False, False),
    # LABA/GCI
    "salmeterol + fluticasona":             ("LABA/GCI",     "50/500 mcg c/12h",       "50/500 mcg c/12h",  True,  False),
    "formoterol + budesonida":              ("LABA/GCI",     "9/320 mcg c/12h",        "36/1.280 mcg/24h",  True,  False),
    "formoterol + beclometasona":           ("LABA/GCI",     "12/200 mcg c/12h",       "12/400 mcg c/12h",  False, False),
    "vilanterol + fluticasona":             ("LABA/GCI",     "22/92 mcg c/24h",        "184/22 mcg c/24h",  False, False),
    "vilanterol + furoato de fluticasona":  ("LABA/GCI",     "22/92 mcg c/24h",        "184/22 mcg c/24h",  False, False),
    # LAMA/LABA/GCI
    "glicopirronio + formoterol + beclometasona": ("LAMA/LABA/GCI","18/10/174 mcg c/12h","18/10/344 mcg c/12h",True, False),
    "umeclidinio + vilanterol + fluticasona":      ("LAMA/LABA/GCI","55/22/92 mcg c/24h", "184/55/22 mcg c/24h",False,False),
    "glicopirronio + indacaterol + mometasona":    ("LAMA/LABA/GCI","114/46/136 mcg c/24h","114/46/136 mcg c/24h",True,False),
    "glicopirronio + formoterol + budesonida":     ("LAMA/LABA/GCI","14.4/10/320 mcg c/12h","14.4/10/320 mcg c/12h",False,False),
}

def inferir_vtm(vtm_nom: str):
    """Cerca el VTM al mapa i retorna (classe, dosi, dosi_max, phf, matma) o None."""
    if not vtm_nom:
        return None
    v = vtm_nom.lower().strip()
    # Cerca exacta
    if v in VTM_MAP:
        return VTM_MAP[v]
    # Cerca parcial (el VTM pot tenir variacions)
    for key, val in VTM_MAP.items():
        if key in v or v in key:
            return val
    return None

# ── Descarregar fitxers de GitHub ─────────────────────────────────────────────
def download_from_github(fname: str, mode: str = "binary") -> str:
    url = f"https://raw.githubusercontent.com/{GITHUB_OWNER}/{GITHUB_REPO}/main/{fname}"
    headers = {"Authorization": f"token {GITHUB_TOKEN}"} if GITHUB_TOKEN else {}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    local_path = f"/tmp/{fname}"
    if mode == "text":
        with open(local_path, "w", encoding="utf-8") as f:
            f.write(r.text)
    else:
        with open(local_path, "wb") as f:
            f.write(r.content)
    print(f"  Descarregat {fname} ({len(r.content)/1024:.1f} KB)")
    return local_path

# ── MODE: --detecta ───────────────────────────────────────────────────────────
def mode_detecta():
    print("\n=== MODE DETECTA ===")

    print("1. Descarregant Excel...")
    excel_path = download_from_github(EXCEL_FNAME, mode="binary")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Recollir CNs existents (col D, índex 3)
    cns_existents = set()
    skip = True
    for row in ws.iter_rows(values_only=True):
        if skip:
            skip = False
            continue
        cn_cell = str(row[3] or "").strip()
        for cn in re.split(r"[,\n\s]+", cn_cell):
            cn = cn.strip()
            if cn and cn != "0":
                cns_existents.add(cn)
    print(f"  CNs existents: {len(cns_existents)}")

    print("2. Consultant CIMA...")
    medicaments_cima = []
    pagina = 1
    while True:
        params = {**CIMA_PARAMS, "pagina": pagina}
        r = requests.get(CIMA_URL, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        items = data.get("resultados", [])
        if not items:
            break
        medicaments_cima.extend(items)
        if len(items) < 100:
            break
        pagina += 1
        time.sleep(0.5)
    print(f"  Total CIMA: {len(medicaments_cima)}")

    # Detectar novetats per CN
    novetats_raw = []
    for med in medicaments_cima:
        cn = str(med.get("cn", "")).strip().zfill(6)
        if cn and cn != "000000" and cn not in cns_existents:
            novetats_raw.append(med)
    print(f"  CNs nous detectats: {len(novetats_raw)}")

    # Filtrar i enriquir
    today = date.today().isoformat()
    afegits = 0
    descartats = 0

    for med in novetats_raw:
        cn   = str(med.get("cn", "")).strip().zfill(6)
        nom  = med.get("nombre", "").strip()
        vtm  = med.get("vtm", {}) or {}
        ia   = vtm.get("nombre", "").strip() or nom
        nreg = med.get("nregistro", "")
        cima_url = f"https://cima.aemps.es/cima/publico/detalle.html?nregistro={nreg}"
        forma = (med.get("formaFarmaFull", "") or med.get("formaFarma", "")).strip()

        # Filtre: forma farmacèutica ha de ser inhalatòria
        if not es_forma_inhalatoria(forma):
            descartats += 1
            continue

        # Inferir dispositiu (inclou link Scientia Salut)
        tipus, co2, flux, maniobra, link_disp = inferir_dispositiu(nom, forma)

        # Inferir classe, dosi i PHF des del VTM
        vtm_info = inferir_vtm(ia)
        if vtm_info:
            cat, dosi, dosi_max, phf, matma = vtm_info
            dosi_text = f"{dosi}\n(D. màx. {dosi_max})"
        else:
            cat      = ""   # Tu ho ompliràs
            dosi_text = ""  # Tu ho ompliràs
            phf      = False
            matma    = False

        ws.append([
            cat,        # A Classe terapèutica
            ia,         # B Principi actiu
            nom,        # C Nom comercial
            cn,         # D CN
            forma,      # E Dispositiu/Presentació
            dosi_text,  # F Dosi recomanada
            tipus,      # G Tipus dispositiu
            co2,        # H Petjada CO₂
            flux,       # I Flux inspiratori
            maniobra,   # J Maniobra
            link_disp,  # K Link instruccions (Scientia Salut)
            today,      # L Data incorporació
            cima_url,   # M Origen CIMA
            "NOU — PENDENT",  # N Estat validació
            "★" if phf else "",   # O PHF
            "MATMA" if matma else "",  # P MATMA
        ])
        afegits += 1

    wb.save(EXCEL_FNAME)
    with open("novetats_count.txt", "w") as f:
        f.write(str(afegits))
    print(f"✅ DETECTA: {afegits} novetats | {descartats} descartades (no inhalatòries).")

# ── Llegir novetats validades ──────────────────────────────────────────────────
def llegir_novetats_validades(excel_path: str) -> list:
    """
    Retorna NOMÉS files novetats validades:
      - Col N buida (validat per tu)
      - Col M conté URL cima.aemps.es
      - Col L té data
    Els 96 originals NO es toquen (no tenen URL CIMA ni data).
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    novetats = []
    skip = True
    for row in ws.iter_rows(values_only=True):
        if skip:
            skip = False
            continue
        if len(row) <= 13:
            continue
        estat  = str(row[13] or "").strip()
        origen = str(row[12] or "").strip()
        data_l = row[11]
        if estat not in ("", "None"):
            continue
        if "cima.aemps.es" not in origen:
            continue
        if not data_l:
            continue
        cat   = str(row[0] or "").strip()
        ia    = str(row[1] or "").strip()
        if not cat or not ia:
            continue
        brand = str(row[2] or "").strip()
        disp  = str(row[4] or "").strip()
        cn    = str(row[3] or "").strip()
        link  = str(row[10] or "").strip()
        phf   = row[14]
        matma = row[15]
        dosi_f= str(row[5] or "").strip()
        tipus = str(row[6] or "").strip()
        starred = bool(phf and str(phf).strip() not in ("","None","nan","0","False"))
        matma_b = bool(matma and str(matma).strip() not in ("","None","nan","0","False",""))
        cns = [c.strip() for c in re.split(r"[,\n]", cn) if c.strip()]
        farmac = {
            "cat": cat, "ia": ia, "brands": brand, "disp": disp,
            "cn": ", ".join(cns), "tipus": tipus,
            "link": link if link.startswith("http") else "",
            "starred": starred, "matma": matma_b,
        }
        # Dosi
        farmac.update(parsear_dosi(dosi_f, cat))
        novetats.append(farmac)
    wb.close()
    print(f"  Novetats validades: {len(novetats)}")
    return novetats

def parsear_dosi(dosi_text: str, cat: str) -> dict:
    if not dosi_text or str(dosi_text).strip() in ("", "nan"):
        return {}
    text = str(dosi_text).strip()
    result = {}
    dmx_match = re.search(r"[Dd]\.?\s*[Mm]à?x[\.:]?\s*(.+?)(?:\n|$)", text)
    if dmx_match:
        result["dmx"] = dmx_match.group(1).strip()
    if "MPOC" in text and ("Asma" in text or "asma" in text):
        mpoc_m = re.search(r"MPOC\s*[:\-]?\s*(.+?)(?=Asma|$)", text, re.IGNORECASE|re.DOTALL)
        if mpoc_m: result["mpoc"] = mpoc_m.group(1).strip().split("\n")[0].strip()
        baixa_m = re.search(r"[Bb]aixa\s*[:\-]?\s*(.+?)(?=Mitjana|Alta|D\.?\s*[Mm]à?x|$)", text, re.IGNORECASE|re.DOTALL)
        mitj_m  = re.search(r"Mitjana\s*[:\-]?\s*(.+?)(?=Alta|D\.?\s*[Mm]à?x|$)", text, re.IGNORECASE|re.DOTALL)
        alta_m  = re.search(r"Alta\s*[:\-]?\s*(.+?)(?=D\.?\s*[Mm]à?x|$)", text, re.IGNORECASE|re.DOTALL)
        if baixa_m: result["asma_baixa"]  = baixa_m.group(1).strip().split("\n")[0].strip()
        if mitj_m:  result["asma_mitjana"] = mitj_m.group(1).strip().split("\n")[0].strip()
        if alta_m:  result["asma_alta"]    = alta_m.group(1).strip().split("\n")[0].strip()
    else:
        dosi_neta = re.sub(r"D\.?\s*[Mm]à?x[\.:]?.+", "", text).strip()
        result["dose"] = dosi_neta
    return result

def farmac_a_js(d: dict) -> str:
    parts = []
    parts.append(f"cat:{json.dumps(d['cat'], ensure_ascii=False)}")
    parts.append(f"ia:{json.dumps(d['ia'], ensure_ascii=False)}")
    if d.get("starred"): parts.append("starred:true")
    if d.get("matma"):   parts.append("matma:true")
    parts.append(f"brands:{json.dumps(d['brands'], ensure_ascii=False)}")
    parts.append(f"disp:{json.dumps(d['disp'], ensure_ascii=False)}")
    for camp in ("dose","mpoc","asma_baixa","asma_mitjana","asma_alta"):
        if d.get(camp): parts.append(f"{camp}:{json.dumps(d[camp], ensure_ascii=False)}")
    if d.get("dmx"):  parts.append(f"dmx:{json.dumps(d['dmx'], ensure_ascii=False)}")
    parts.append(f"tipus:{json.dumps(d['tipus'], ensure_ascii=False)}")
    if d.get("link"): parts.append(f"link:{json.dumps(d['link'], ensure_ascii=False)}")
    return "  {" + ",\n   ".join(parts) + "}"

def fusionar_catalog(html_path: str, novetats: list) -> str:
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()
    pattern = re.compile(r"(var catalog=\[)(.*?)(\];)", re.DOTALL)
    match = pattern.search(html)
    if not match:
        raise ValueError("No s'ha trobat 'var catalog=[...];' al HTML.")
    catalog_actual = match.group(2).strip()
    n_actuals = catalog_actual.count("{cat:")
    print(f"  Fàrmacs al HTML: {n_actuals} | Novetats: {len(novetats)}")
    novetats_js = ",\n" + ",\n".join([farmac_a_js(f) for f in novetats]) if novetats else ""
    nou_bloc = f"var catalog=[{catalog_actual}{novetats_js}];"
    html_nou = html[:match.start()] + nou_bloc + html[match.end():]
    print(f"  ✅ Total: {n_actuals + len(novetats)} fàrmacs")
    return html_nou

def mode_regenera():
    print("\n=== MODE REGENERA ===")
    excel_path = download_from_github(EXCEL_FNAME, mode="binary")
    novetats = llegir_novetats_validades(excel_path)
    if not novetats:
        print("  ℹ️  Cap novetat validada.")
        with open("publicar_count.txt", "w") as f: f.write("0")
        return
    html_path = download_from_github(HTML_FNAME, mode="text")
    html_nou = fusionar_catalog(html_path, novetats)
    with open(HTML_FNAME, "w", encoding="utf-8") as f:
        f.write(html_nou)
    print(f"  ✅ {HTML_FNAME} desat.")
    with open("publicar_count.txt", "w") as f:
        f.write(str(len(novetats)))

def mode_comprova_pendents():
    print("\n=== COMPROVA PENDENTS ===")
    excel_path = download_from_github(EXCEL_FNAME, mode="binary")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    skip = True
    for row in ws.iter_rows(values_only=True):
        if skip: skip=False; continue
        estat = str(row[13] or "").strip() if len(row)>13 else ""
        if "NOU" in estat and "PENDENT" in estat: count += 1
    wb.close()
    with open("pendents_count.txt","w") as f: f.write(str(count))
    print(f"✅ {count} pendents.")

def mode_comprova_publicar():
    print("\n=== COMPROVA PUBLICAR ===")
    excel_path = download_from_github(EXCEL_FNAME, mode="binary")
    novetats = llegir_novetats_validades(excel_path)
    with open("publicar_count.txt","w") as f: f.write(str(len(novetats)))
    print(f"✅ {len(novetats)} llestos per publicar.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="RespirIA CIMA Updater v7.0")
    parser.add_argument("--mode", choices=["detecta","comprova-pendents",
                        "comprova-publicar","regenera","tot"], required=True)
    args = parser.parse_args()
    if args.mode == "detecta":          mode_detecta()
    elif args.mode == "comprova-pendents": mode_comprova_pendents()
    elif args.mode == "comprova-publicar": mode_comprova_publicar()
    elif args.mode == "regenera":       mode_regenera()
    elif args.mode == "tot":            mode_detecta(); mode_regenera()
